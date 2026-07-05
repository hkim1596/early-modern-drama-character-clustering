"""Stage 01 (unified) — XML extraction, speaker-map summary, character table.

Replaces the former 01_extract_characters.py, 01c_build_play_merge_summary.py,
01d_apply_verify_flags.py and 01e_build_character_table.py. All bookkeeping
lives in ONE workbook, data/corpus_master.xlsx:

  plays         one row per catalogue entry (statuses, per-play speaker stats,
                notes, verify_flag/verify_note) — the old merge summary
  speaker_maps  THE source of mapping truth: one row per (source_file, bucket,
                key, form); bucket = speaker | collective | unresolved.
                Edit here, then rerun `summary` + `table`.
  extraction    one row per XML file from the last extract run
  characters    one row per character bucket with `status` =
                kept | dropped_short | unresolved_excluded  (replaces the old
                dropped_short_characters.csv and unresolved_excluded.csv)
  decisions     round-2 adjudication log (static record)
  queue         round-2 triage queue with final statuses (static record)

The only outputs kept outside the workbook:
  data/character_json_single_plays/*.json   per-play extraction cache
  data/character_table.csv                  characters WITH full speech text
                                            (text exceeds Excel's 32,767-char
                                            cell limit, so it cannot live in
                                            the workbook); consumed by stage 02

Usage::

    python code/01_build_corpus.py                  # all three stages
    python code/01_build_corpus.py --stage extract
    python code/01_build_corpus.py --stage summary  # also re-applies verify flags
    python code/01_build_corpus.py --stage table

Unresolved prefix buckets are excluded from character_table.csv (and thus
from clustering) unless config.INCLUDE_UNRESOLVED_PREFIXES is True; they are
always listed on the `characters` sheet with status `unresolved_excluded`.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from bs4 import BeautifulSoup
except ImportError:  # pragma: no cover
    sys.exit("This script needs BeautifulSoup:  pip install beautifulsoup4")

_HERE = Path(__file__).resolve().parent
_ROOT = _HERE.parent
sys.path.insert(0, str(_HERE))
import config  # noqa: E402

XML_DIR = _ROOT / "data" / "tcp_drama_single_plays"
JSON_DIR = _ROOT / "data" / "character_json_single_plays"
CATALOGUE = _ROOT / "data" / "deep_tcp_british_drama.xlsx"
MASTER = _ROOT / "data" / "corpus_master.xlsx"
TABLE_ORIG = _ROOT / "data" / "character_table_original.csv"
TABLE_MOD = _ROOT / "data" / "character_table_modern.csv"

_WS_RE = re.compile(r"\s+")

F_HEAD = Font(name="Arial", bold=True)
F_BODY = Font(name="Arial", size=10)
AL_TOP = Alignment(vertical="top")
FILL_HEAD = PatternFill("solid", start_color="DDDDDD")


# ---------------------------------------------------------------- helpers --
def collapse_ws(text: str) -> str:
    return _WS_RE.sub(" ", text).strip() if text else ""


def norm_key(s: str) -> str:
    """Whitespace-insensitive prefix key (NFC + strip all whitespace)."""
    return re.sub(r"\s+", "", unicodedata.normalize("NFC", s))


_LONG_S = str.maketrans({"\u017f": "s", "\u0283": "s"})
_GAP_RE = re.compile(r"[\u3008\u3009\u25ca\u2022\u25aa\u2026]")
_VV_RE = re.compile(r"\bVV|\bVv|\bvv")


def normalize_speech_text(text: str) -> str:
    """Embedding-input cleanup: long-s, VV->W, TCP damage glyphs (see config)."""
    text = text.translate(_LONG_S)
    text = _VV_RE.sub(lambda m: "W" if m.group(0)[0] == "V" else "w", text)
    text = _GAP_RE.sub(" ", text)
    return collapse_ws(text)


class SpellingModernizer:
    """Rule-based EME spelling modernization using MorphAdorner data.

    Lookup order per token (case restored afterwards):
      1. exact-case hit in ememergedspellingpairs.tab   (Ile -> I'll)
      2. lowercased hit in the same map                 (Loue -> love)
      3. u/v, i/j, ie->y heuristic candidates, accepted only if the candidate
         itself validates against the pair map or standardspellings.txt
      4. otherwise the token is left unchanged
    The pair map is consulted BEFORE any "already standard" early exit because
    standardspellings.txt retains archaic-but-valid forms (wee, doe, hee).
    """

    OVERRIDES = {"shalbe": "shall be", "shalbee": "shall be",
                 "wilbe": "will be", "wilbee": "will be",
                 # bare tis in drama = 'tis (it is), not 'this'
                 "tis": "'tis"}
    # leading/trailing apostrophes are part of the token: th' -> the, 'tis -> it's
    _TOK = re.compile(r"(['\u2019]?[^\W\d_]+(?:['\u2019-][^\W\d_]+)*['\u2019]?)",
                      re.UNICODE)

    def __init__(self, data_dir: Path):
        self.pairs: dict[str, str] = {}
        for line in (data_dir / "ememergedspellingpairs.tab").open(encoding="utf-8-sig"):
            parts = line.rstrip("\n").split("\t")
            if len(parts) != 2:
                continue
            v, s = parts
            if any(ch in v or ch in s for ch in ("^", "\u204f", "\u2022")):
                continue
            self.pairs[v] = s
        self.std = {l.strip() for l in
                    (data_dir / "standardspellings.txt").open(encoding="utf-8-sig")}
        self.stats = Counter()
        self._memo: dict[str, str] = {}

    @staticmethod
    def _restore_case(orig: str, repl: str) -> str:
        alpha = next((c for c in orig if c.isalpha()), "")
        stripped = [c for c in orig if c.isalpha()]
        if stripped and all(c.isupper() for c in stripped) and len(stripped) > 1:
            return repl.upper()
        if alpha.isupper():
            for i, c in enumerate(repl):
                if c.isalpha():
                    return repl[:i] + c.upper() + repl[i + 1:]
        return repl

    def _candidates(self, w: str):
        yield re.sub(r"^v", "u", w)
        yield re.sub(r"^u", "v", w)
        yield re.sub(r"(?<=[a-z])u(?=[aeiou])", "v", w)
        yield re.sub(r"(?<=[aeiou])v(?=[a-z])", "u", w)
        yield re.sub(r"^i(?=[aeiou])", "j", w)
        yield re.sub(r"ie$", "y", w)

    def word(self, w: str) -> str:
        hit = self._memo.get(w)
        if hit is not None:
            return hit
        out = self._word(w)
        self._memo[w] = out
        return out

    def _word(self, w: str) -> str:
        key = w.replace("\u2019", "'")
        low = key.lower()
        if low in self.OVERRIDES:
            self.stats["pair"] += 1
            return self._restore_case(w, self.OVERRIDES[low])
        hit = self.pairs.get(key)
        if hit is None:
            hit = self.pairs.get(low)
            if hit is not None:
                hit = self._restore_case(w, hit)
        if hit is not None:
            self.stats["pair" if hit != w else "identity"] += 1
            return hit
        if low in self.std:
            self.stats["standard"] += 1
            return w
        for cand in self._candidates(low):
            if cand == low:
                continue
            if cand in self.pairs:
                self.stats["heuristic"] += 1
                return self._restore_case(w, self.pairs[cand])
            if cand in self.std:
                self.stats["heuristic"] += 1
                return self._restore_case(w, cand)
        self.stats["oov"] += 1
        return w

    def text(self, text: str) -> str:
        return self._TOK.sub(lambda m: self.word(m.group(0)), text)


def fmt(pairs):
    return "; ".join(f"{k} ({v})" for k, v in pairs)


def wsum(s: str | None) -> int:
    return sum(int(m) for m in re.findall(r"\((\d+)\)", s or ""))


def open_master() -> openpyxl.Workbook:
    if MASTER.exists():
        return openpyxl.load_workbook(MASTER)
    return openpyxl.Workbook()


def replace_sheet(wb: openpyxl.Workbook, name: str):
    """Delete-and-recreate a sheet, keeping its position; return the sheet."""
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        del wb[name]
        ws = wb.create_sheet(name, idx)
    else:
        ws = wb.create_sheet(name)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    return ws


def write_header(ws, headers, widths=None):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font, c.fill = F_HEAD, FILL_HEAD
        if widths:
            ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
    ws.freeze_panes = "A2"


def save_master(wb) -> None:
    import shutil, tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)  # write locally, then copy once (cloud-synced data dir)
    shutil.copy(tmp_path, MASTER)


def load_maps(wb) -> dict[str, dict]:
    """speaker_maps sheet -> {source_file: {speakers:{k:[f,..]}, collective:[], unresolved:[]}}."""
    ws = wb["speaker_maps"]
    maps: dict[str, dict] = {}
    for sf, bucket, key, form in ws.iter_rows(min_row=2, values_only=True):
        if not sf:
            continue
        e = maps.setdefault(sf, {"speakers": {}, "collective": [], "unresolved": []})
        if bucket == "speaker":
            e["speakers"].setdefault(key, []).append(form)
        elif bucket == "collective":
            e["collective"].append(form)
        elif bucket == "unresolved":
            e["unresolved"].append(form)
    return maps


# ---------------------------------------------------------------- extract --
def extract_speeches(xml_path: Path) -> dict:
    soup = BeautifulSoup(xml_path.read_text(encoding="utf-8", errors="ignore"),
                         "html.parser")
    speeches: dict[str, list[str]] = {}
    order: list[str] = []
    current = None
    sp_elements = soup.find_all("sp")
    n_no_speaker = n_leading = 0
    for sp in sp_elements:
        el = sp.find("speaker")
        if el is not None:
            name = collapse_ws(el.get_text(" "))
            if name:
                current = name
        else:
            n_no_speaker += 1
        if current is None:
            n_leading += 1
            continue
        body = collapse_ws(" ".join(e.get_text(" ") for e in sp.find_all(["l", "p"])))
        if not body:
            continue
        if current not in speeches:
            speeches[current] = []
            order.append(current)
        speeches[current].append(body)
    return {
        "speeches": {n: collapse_ws(" ".join(speeches[n])) for n in order},
        "n_sp": len(sp_elements),
        "n_sp_without_speaker": n_no_speaker,
        "n_leading_skipped": n_leading,
    }


def stage_extract(wb) -> None:
    JSON_DIR.mkdir(parents=True, exist_ok=True)
    xml_files = sorted(p for p in XML_DIR.rglob("*.xml") if "_unedited" not in p.stem)
    print(f"extract: {len(xml_files)} XML files (skipping *_unedited)")
    rows, written = [], 0
    for xp in xml_files:
        ex = extract_speeches(xp)
        stem = xp.stem
        tcp, _, part = stem.partition(".")
        wc = {n: len(t.split()) for n, t in ex["speeches"].items()}
        rows.append([xp.name, tcp, part, len(wc), sum(wc.values()), ex["n_sp"],
                     ex["n_sp_without_speaker"], ex["n_leading_skipped"],
                     "written" if wc else "skipped_no_characters",
                     "; ".join(f"{n} ({c})" for n, c in wc.items())])
        if not wc:
            continue
        record = {
            "tcp": tcp, "part": part, "source_file": xp.name,
            "n_characters": len(wc), "n_words": sum(wc.values()),
            "characters": list(ex["speeches"].keys()),
            "word_counts": wc, "speeches": ex["speeches"],
        }
        (JSON_DIR / f"{stem}.json").write_text(
            json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        written += 1
    ws = replace_sheet(wb, "extraction")
    hdr = ["source_file", "tcp", "part", "n_characters", "n_words", "n_sp",
           "n_sp_without_speaker", "n_leading_skipped", "status", "speakers"]
    write_header(ws, hdr, [14, 9, 5, 11, 9, 7, 9, 9, 20, 60])
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = F_BODY
    print(f"extract: {written} JSONs written; extraction sheet {len(rows)} rows")


# ---------------------------------------------------------------- summary --
STATIC_NOTES = {
    "A04632.19.xml": "FIXED 2026-07-05: re-split; full Masque of Queens text restored "
                     "(it has no speech prefixes); duplicated Barriers tail trimmed.",
    "A04632.24.xml": "FIXED 2026-07-05: Challenge at Tilt wholly lost to untranscribed "
                     "pages (missing-page stub); surviving Irish Masque text moved to .25.",
    "A04632.25.xml": "FIXED 2026-07-05: Irish Masque dialogue relocated here from the "
                     ".24 slot; opening lost to missing pages.",
}

PLAYS_HEADERS = ["tcp", "part", "source_file", "title", "status", "n_words",
                 "n_speakers", "standardized_speakers", "prefix_variants",
                 "original_prefixes", "collective_prefixes", "unresolved_prefixes",
                 "notes", "roles", "Stage Directions and Speech Prefixes",
                 "verify_flag", "verify_note"]


def stage_summary(wb) -> None:
    mappings = load_maps(wb)

    # preserve notes edited on the plays sheet across rebuilds
    old_notes: dict[str, str] = {}
    if "plays" in wb.sheetnames:
        po = wb["plays"]
        h = [c.value for c in po[1]]
        if "source_file" in h and "notes" in h:
            i_sf, i_no = h.index("source_file"), h.index("notes")
            for r in po.iter_rows(min_row=2, values_only=True):
                if r[i_sf]:
                    old_notes[r[i_sf]] = r[i_no] or ""

    ref = openpyxl.load_workbook(CATALOGUE, read_only=True).active
    rrows = list(ref.iter_rows(values_only=True))
    h = list(rrows[0])
    ci = {k: h.index(k) for k in ("TCP", "record_type", "roles",
                                  "Stage Directions and Speech Prefixes", "item_title")}
    ref_rows = [r for r in rrows[1:] if r[ci["TCP"]] and str(r[ci["TCP"]]).strip()]
    ref_ids = {str(r[ci["TCP"]]).strip() for r in ref_rows}

    exw = wb["extraction"]
    eh = [c.value for c in exw[1]]
    extraction = {}
    for r in exw.iter_rows(min_row=2, values_only=True):
        d = dict(zip(eh, r))
        if d["source_file"]:
            extraction[d["source_file"][:-4]] = d
    jsons = {p.stem for p in JSON_DIR.glob("*.json")}
    consumed_extraction: set[str] = set()

    rows_out, problems = [], []
    for x in ref_rows:
        tcp_id = str(x[ci["TCP"]]).strip()
        rtype = str(x[ci["record_type"]] or "")
        base = {
            "title": str(x[ci["item_title"]] or ""),
            "roles": str(x[ci["roles"]] or ""),
            "Stage Directions and Speech Prefixes":
                str(x[ci["Stage Directions and Speech Prefixes"]] or ""),
            "n_speakers": "", "standardized_speakers": "", "prefix_variants": "",
            "original_prefixes": "", "collective_prefixes": "",
            "unresolved_prefixes": "", "notes": "",
        }
        stem = tcp_id if tcp_id in extraction else None
        if stem is None and tcp_id != "missing" \
                and f"{tcp_id}.1" in extraction and f"{tcp_id}.1" not in ref_ids:
            stem = f"{tcp_id}.1"
        ex = extraction.get(stem) if stem else None
        if stem:
            consumed_extraction.add(stem)

        if tcp_id == "missing":
            base.update(tcp="missing", part="", source_file="",
                        status="missing_from_tcp", n_words="")
            rows_out.append(base); continue
        if ex is None:
            if rtype == "Collection":
                base.update(tcp=tcp_id, part="", source_file="",
                            status="collection_volume", n_words="")
            else:
                base.update(tcp=tcp_id, part="", source_file="",
                            status="no_xml", n_words="")
                problems.append((tcp_id, "-", "-", "catalogue play has no extraction"))
            rows_out.append(base); continue
        if stem not in jsons:
            base.update(tcp=ex["tcp"], part=ex["part"], source_file=ex["source_file"],
                        status="no_speaker_text", n_words=0, n_speakers=0)
            rows_out.append(base); continue

        j = json.loads((JSON_DIR / f"{stem}.json").read_text(encoding="utf-8"))
        wc: dict[str, int] = j["word_counts"]
        row = dict(base)
        row.update(tcp=j["tcp"], part=j["part"], source_file=j["source_file"],
                   n_words=j["n_words"],
                   original_prefixes=fmt(sorted(wc.items(), key=lambda kv: -kv[1])),
                   notes=old_notes.get(j["source_file"], ""))
        m = mappings.get(j["source_file"])
        if not m:
            row.update(status="pending")
            rows_out.append(row); continue

        lookup = dict(wc)
        by_norm = {norm_key(k): k for k in wc}
        consumed: dict[str, str] = {}

        def take(prefix, dest):
            key = prefix if prefix in lookup else by_norm.get(norm_key(prefix))
            if key is None or key in consumed:
                problems.append((j["source_file"], dest, prefix,
                                 "phantom" if key is None else "duplicate"))
                return None
            consumed[key] = dest
            return key

        speakers, variants = [], []
        for name, forms in m["speakers"].items():
            got = [k for f in forms if (k := take(f, name)) is not None]
            if not got:
                problems.append((j["source_file"], name, "-", "no prefixes matched"))
                continue
            got.sort(key=lambda k: -wc[k])
            speakers.append((name, sum(wc[k] for k in got)))
            variants.append(got[0] if len(got) == 1
                            else f"{got[0]} ({'; '.join(got[1:])})")
        coll = [k for f in m["collective"] if (k := take(f, "~collective"))]
        unres = [k for f in m["unresolved"] if (k := take(f, "~unresolved"))]
        auto = [k for k in wc if k not in consumed]
        unres += auto
        assert sum(t for _, t in speakers) + sum(wc[k] for k in coll) \
            + sum(wc[k] for k in unres) == j["n_words"], j["source_file"]

        order = sorted(range(len(speakers)), key=lambda i: -speakers[i][1])
        row.update(
            status="llm_round1", n_speakers=len(speakers),
            standardized_speakers=fmt([speakers[i] for i in order]),
            prefix_variants=" | ".join(f"{speakers[i][0]}: {variants[i]}"
                                       for i in order),
            collective_prefixes=fmt(sorted(((k, wc[k]) for k in coll),
                                           key=lambda kv: -kv[1])),
            unresolved_prefixes=fmt(sorted(((k, wc[k]) for k in unres),
                                           key=lambda kv: -kv[1])) +
                                (" [+auto: " + "; ".join(auto) + "]" if auto else ""),
        )
        rows_out.append(row)

    assert len(rows_out) == len(ref_rows), (len(rows_out), len(ref_rows))
    for o in sorted(set(extraction) - consumed_extraction):
        problems.append((o, "-", "-", "extraction file not in catalogue TCP column"))

    # verify flags (former 01d), computed inline
    n_flags = 0
    for row in rows_out:
        sf, flag, note = row.get("source_file", ""), None, None
        if sf in STATIC_NOTES:
            flag, note = "NOTE", STATIC_NOTES[sf]
        elif row["status"] == "llm_round1":
            nw = row["n_words"] or 0
            un = wsum(row["unresolved_prefixes"])
            if nw and un / nw >= 0.05:
                flag = "HIGH_UNRESOLVED"
                note = (f"{100 * un / nw:.1f}% of words unresolved ({un:,}/{nw:,}) — "
                        "round-2 adjudication target (see notes col).")
                if sf == "A01509.xml":
                    flag = "HIGH_UNRESOLVED; TEXT_LOSS"
                    note += (" ALSO 132 of 317 speeches have no <speaker> prefix — "
                             "most unattributed text in corpus.")
        row["verify_flag"], row["verify_note"] = flag or "", note or ""
        n_flags += bool(flag)

    ws = replace_sheet(wb, "plays")
    widths = [9, 5, 13, 30, 15, 9, 10, 46, 46, 46, 22, 26, 40, 40, 40, 18, 70]
    write_header(ws, PLAYS_HEADERS, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PLAYS_HEADERS))}{len(rows_out) + 1}"
    fills = {"llm_round1": PatternFill("solid", start_color="E2EFDA"),
             "pending": PatternFill("solid", start_color="FFF2CC"),
             "no_speaker_text": PatternFill("solid", start_color="EDEDED"),
             "collection_volume": PatternFill("solid", start_color="D9E1F2"),
             "missing_from_tcp": PatternFill("solid", start_color="FCE4D6"),
             "no_xml": PatternFill("solid", start_color="FFC7CE")}
    amber = PatternFill("solid", fgColor="FFEB9C")
    blue = PatternFill("solid", fgColor="DDEBF7")
    for i, row in enumerate(rows_out, 2):
        for j, hname in enumerate(PLAYS_HEADERS, 1):
            c = ws.cell(row=i, column=j, value=row[hname])
            c.font, c.alignment = F_BODY, AL_TOP
        if row["status"] in fills:
            ws.cell(row=i, column=5).fill = fills[row["status"]]
        if row["verify_flag"]:
            ws.cell(row=i, column=16).fill = \
                blue if row["verify_flag"] == "NOTE" else amber

    counts = Counter(r["status"] for r in rows_out)
    print(f"summary: {len(rows_out)} rows (= catalogue rows with TCP filled)")
    for k in ("llm_round1", "pending", "no_speaker_text",
              "collection_volume", "missing_from_tcp", "no_xml"):
        if counts.get(k):
            print(f"  {k}: {counts[k]}")
    print(f"  verify flags: {n_flags}")
    if problems:
        print(f"summary: PROBLEMS ({len(problems)}):")
        for p in problems:
            print("  ", p)


# ------------------------------------------------------------------ table --
def stage_table(wb) -> None:
    mappings = load_maps(wb)
    modernizer = SpellingModernizer(config.MORPHADORNER_DATA)
    rows, sheet_rows = [], []
    for sf, m in sorted(mappings.items()):
        stem = sf[:-4] if sf.endswith(".xml") else sf
        jp = JSON_DIR / f"{stem}.json"
        if not jp.exists():
            continue
        pj = json.loads(jp.read_text(encoding="utf-8"))
        speeches, counts = pj["speeches"], pj["word_counts"]
        by_norm: dict[str, list[str]] = {}
        for jf in speeches:
            by_norm.setdefault(norm_key(jf), []).append(jf)

        def bucket(name, forms):
            texts, found = [], []
            for f in forms:
                for jf in ([f] if f in speeches else by_norm.get(norm_key(f), [])):
                    texts.append(speeches[jf])
                    found.append(jf)
            orig = " ".join(texts)
            modern = modernizer.text(normalize_speech_text(orig))
            return {"TCP": stem, "normalized_name": name,
                    "raw_names": " | ".join(found), "n_forms": len(found),
                    "n_words": len(orig.split()), "n_chars": len(orig),
                    "match_score": "", "speech_text": orig,
                    "_modern": modern}

        buckets = [bucket(k, forms) for k, forms in m["speakers"].items()]
        if m["collective"]:
            buckets.append(bucket("__crowd__", m["collective"]))
        buckets = [b for b in buckets if b["n_forms"]]
        for b in buckets:
            if b["n_chars"] < config.MIN_SPEECH_CHARS:
                status = "dropped_short"
            else:
                status = "kept"
                rows.append(b)
            sheet_rows.append([b["TCP"], b["normalized_name"], status,
                               b["n_forms"], b["n_words"], b["n_chars"],
                               b["raw_names"]])
        for form in m["unresolved"]:
            b = bucket(f"[unresolved] {form}", [form])
            if not b["n_forms"]:
                continue
            if config.INCLUDE_UNRESOLVED_PREFIXES:
                status = ("kept" if b["n_chars"] >= config.MIN_SPEECH_CHARS
                          else "dropped_short")
                if status == "kept":
                    rows.append(b)
            else:
                status = "unresolved_excluded"
            sheet_rows.append([b["TCP"], b["normalized_name"], status,
                               b["n_forms"], b["n_words"], b["n_chars"],
                               b["raw_names"]])

    import csv as _csv
    fields = ["TCP", "normalized_name", "raw_names", "n_forms",
              "n_words", "n_chars", "match_score", "speech_text"]
    with TABLE_ORIG.open("w", newline="", encoding="utf-8-sig") as fh:
        w = _csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    with TABLE_MOD.open("w", newline="", encoding="utf-8-sig") as fh:
        w = _csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            m = dict(r, speech_text=r["_modern"],
                     n_words=len(r["_modern"].split()),
                     n_chars=len(r["_modern"]))
            w.writerow(m)

    ws = replace_sheet(wb, "characters")
    write_header(ws, ["TCP", "name", "status", "n_forms", "n_words", "n_chars",
                      "forms"], [11, 28, 20, 8, 9, 9, 60])
    ws.auto_filter.ref = f"A1:G{len(sheet_rows) + 1}"
    fills = {"dropped_short": PatternFill("solid", start_color="EDEDED"),
             "unresolved_excluded": PatternFill("solid", start_color="FFF2CC")}
    for i, r in enumerate(sheet_rows, 2):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font, c.alignment = F_BODY, AL_TOP
        if r[2] in fills:
            ws.cell(row=i, column=3).fill = fills[r[2]]

    st = Counter(r[2] for r in sheet_rows)
    print(f"table: {TABLE_ORIG.name} + {TABLE_MOD.name}: {len(rows)} rows each, "
          f"{sum(r['n_words'] for r in rows):,} original words")
    print(f"  characters sheet: {len(sheet_rows)} rows — "
          + ", ".join(f"{k} {v}" for k, v in st.most_common()))
    ms = modernizer.stats
    tot = sum(ms.values()) or 1
    print("  modernization: " + ", ".join(
        f"{k} {v:,} ({100*v/tot:.1f}%)" for k, v in ms.most_common()))


# ------------------------------------------------------------------- main --
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--stage", choices=["extract", "summary", "table", "all"],
                    default="all")
    args = ap.parse_args()

    wb = open_master()
    if args.stage in ("extract", "all"):
        stage_extract(wb)
    if args.stage in ("summary", "all"):
        stage_summary(wb)
    if args.stage in ("table", "all"):
        stage_table(wb)
    save_master(wb)
    print(f"saved: {MASTER.name}")


if __name__ == "__main__":
    main()
